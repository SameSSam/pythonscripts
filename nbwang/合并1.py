
import os#导入os
from openpyxl import load_workbook,Workbook#导入openpyxl
path='/Users/wangwenyan/Desktop/python/输出'
files_list=os.listdir(path)#获取目标文件夹内各文件的名称
print(files_list)
new_wb=Workbook()#新建表格
new_ws=new_wb.active#获取工作簿
# header=['项目','内容','责任人']#表头列表
# new_ws.append(header)#添加新表的表头内容
for file in files_list[1:]:#循环获取各表的名称
    print(file)
    file_path=str(path)+'/'+str(file)#获取表格的路径
   
    wb=load_workbook(file_path)#打开表格
    ws=wb['Sheet1']
    for cells in ws:#循环获取表头以外的表格内容
        new_ws.append(cells.value)#获取的内容写入新表
new_wb.save('/Users/wangwenyan/Desktop/python/输出/完成')#保存表格


