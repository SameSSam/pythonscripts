from datetime import date
import openpyxl as op

def readexcel(file_path):
    return op.load_workbook(file_path)

op1=readexcel('/Users/wangwenyan/Desktop/python/数据文件/42165.xlsx')
op1_1=op1['Sheet1']
op2=readexcel('/Users/wangwenyan/Desktop/python/muban.xlsx')
op2_1=op2['Sheet1']

max_row=op1_1.max_row#最大行数

#发票号码
fphm=op1_1['C2'].value
op2_1['C2']=fphm
op2_1.merge_cells('C2:D2')

#单位名称
dwmc=op1_1['D2'].value
op2_1['C3']=dwmc
op2_1.merge_cells('C3:D3')

#日期
riqi=op1_1['B2'].value
op2_1['G3']=str(riqi)[:11]
op2_1.merge_cells('G3:H3')

#合计
heji=op1_1['J2'].value
op2_1['G19']=heji

if max_row <= 15:
    op2_1.delete_rows(22,20)
    for m in range(1,max_row):
        A=chr(65)
        i='%s%d'%(A,m+4)
        op2_1[i]=m
    B=chr(66)
    C=chr(67)
    D=chr(68)
    E=chr(69)
    F=chr(70)
    for m in range(2,max_row+1):
        mc='%s%d'%(E,m)
        xh='%s%d'%(F,m)
        mc_v=op1_1[mc].value
        xh_v=op1_1[xh].value 
        hwmc='%s%d'%(B,m+3)
        hwxh='%s%d'%(C,m+3)
        op2_1[hwmc]=mc_v
        op2_1[hwxh]=xh_v

    dw='%s%d'%(D,m+3)
    print(mc_v)
    if mc_v=='色带' or '连接线':
        op2_1[dw]='个'
        print('个')
    elif mc_v=='标签机' or '打印机' or '电脑' or '高拍仪' :
        op2_1[dw]='台'
        print('台')
    elif mc_v=='墨粉': 
        op2_1[dw]='瓶'
        print('瓶')
    elif mc_v=='维修费': 
        op2_1[dw]='次'
        print('次')
    else:
        op2_1[dw]='个'
        print('个')
    for m in range(2,max_row+1):
        for n in range(71,74):#chr(97)='a'
            n_1=n-2
            n=chr(n)#ASCII字符
            n_1=chr(n_1)
            i='%s%d'%(n,m)#单元格编号
            i_1='%s%d'%(n_1,m+3)
            cell1=op1_1[i].value#获取data单元格数据
            op2_1[i_1].value=cell1#赋值到test单元格

else:
    op2_1['C23']=fphm
    op2_1.merge_cells('C23:D23')
    op2_1['C24']=dwmc
    op2_1.merge_cells('C24:D24')
    op2_1['G24']=str(riqi)[:11]
    op2_1.merge_cells('G3:H3')
    op2_1['G40']=heji
    op2_1['A19']='小               计'
    for m in range(1,15):
        A=chr(65)
        i='%s%d'%(A,m+4)
        op2_1[i]=m
    for m in range(15,max_row):
        A=chr(65)
        i='%s%d'%(A,m+11)
        op2_1[i]=m
    B=chr(66)
    C=chr(67)
    D=chr(68)
    E=chr(69)
    F=chr(70)
    for m in range (1,16):
        mc='%s%d'%(E,m)
        xh='%s%d'%(F,m)
        mc_v=op1_1[mc].value
        xh_v=op1_1[xh].value 
        hwmc='%s%d'%(B,m+3)
        hwxh='%s%d'%(C,m+3)
        op2_1[hwmc]=mc_v
        op2_1[hwxh]=xh_v
    for m in range (16,max_row+1):
        mc='%s%d'%(E,m)
        xh='%s%d'%(F,m)
        mc_v=op1_1[mc].value
        xh_v=op1_1[xh].value 
        hwmc='%s%d'%(B,m+10)
        hwxh='%s%d'%(C,m+10)
        op2_1[hwmc]=mc_v
        op2_1[hwxh]=xh_v
    for m in range (2,max_row+1):
        if m in range(2,16):
            dw_1=m+3
        if m in range(16,max_row+1):
            dw_1=m+10
        dw='%s%d'%(D,dw_1)
        if mc_v=='色带' or '连接线':
            op2_1[dw]='个'
            print('个')
        elif mc_v=='标签机' or '打印机' or '电脑' or '高拍仪' :
            op2_1[dw]='台'
            print('台')
        elif mc_v=='墨粉': 
            op2_1[dw]='瓶'
            print('瓶')
        elif mc_v=='维修费': 
            op2_1[dw]='次'
            print('次')
        else:
            op2_1[dw]='个'
            print('个')
    for m in range (1,16):
        for n in range(71,74):#chr(97)='a'
            n_1=n-2
            n=chr(n)#ASCII字符
            n_1=chr(n_1)
            i='%s%d'%(n,m)#单元格编号
            i_1='%s%d'%(n_1,m+3)
            cell1=op1_1[i].value#获取data单元格数据
            op2_1[i_1].value=cell1#赋值到test单元格
    for m in range (16,max_row+1):
        for n in range(71,74):#chr(97)='a'
            n_1=n-2
            n=chr(n)#ASCII字符
            n_1=chr(n_1)
            i='%s%d'%(n,m)#单元格编号
            i_1='%s%d'%(n_1,m+10)
            cell1=op1_1[i].value#获取data单元格数据
            op2_1[i_1].value=cell1#赋值到test单元格
    G=chr(71)
    li1=[]
    li2=[]
    for s in range(5,19):
        je='%s%d'%(G,s)
        li1.append(op2_1[je].value)
    op2_1['G19']=sum(li1)
    for s in range(26,max_row+11):
        je='%s%d'%(G,s)
        li2.append(op2_1[je].value)
    op2_1['G39']=sum(li2)
    op2_1['A39']='小               计'
    op2_1.merge_cells('A39:F39')
    
op2.save('/Users/wangwenyan/Desktop/python/输出2.xlsx')

