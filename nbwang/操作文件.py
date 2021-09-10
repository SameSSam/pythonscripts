import os
import openpyxl as op
from os import getgroups, name
import pandas as pd
from pandas.core import indexing
from pandas.core.indexing import convert_to_index_sliceable 
import xlrd
from openpyxl.styles import fills,colors,NamedStyle,Font,Side,Border,PatternFill,Alignment,Protection
import xlsxwriter

def readexcel(filepath, sheet_name):
    return pd.read_excel(filepath, sheet_name = sheet_name)
print('请输入文件名(不带文件类型名称），例如202105')
x=input()
x=str(x)
print('需要处理的sheet，例如Sheet1')
y=input()
y=str(y)
df1 = readexcel('/Users/wangwenyan/Desktop/python/{0}.xlsx'.format(x),'{0}'.format(y))
df2 = df1.groupby('发票号码')
df1['合计'] = df2['未税金额'].transform('sum')
class_list = list(df1['发票号码'].drop_duplicates())
for i in class_list:
    df1_1 = df1[df1['发票号码']==i]

    df1_1.to_excel('/Users/wangwenyan/Desktop/python/数据文件/%s.xlsx'%(i))

import os
import openpyxl as op
import pandas as pd

# op1=readexcel('/Users/wangwenyan/Desktop/python/数据文件/95089.xlsx')
# op1_1=op1['Sheet1']

path1='/Users/wangwenyan/Desktop/python/数据文件'
fileslist1=os.listdir(path1)#获取目标文件夹内各文件的名称

for file in fileslist1:#循环获取各表的名称
    
    file_path=str(path1)+'/'+str(file)#获取表格的路径
    wb=op.load_workbook(file_path)#打开表格
    print(file_path)
    ws=wb['Sheet1']

    max_row=ws.max_row#最大行数


    if max_row <= 15:
        op2=op.load_workbook('/Users/wangwenyan/Desktop/python/muban_1.xlsx')
        op2_1=op2['Sheet1']

        fphm=ws['C2'].value
        op2_1['C2']=fphm
        op2_1.merge_cells('C2:D2')
        dwmc=ws['D2'].value
        op2_1['C3']=dwmc
        op2_1.merge_cells('C3:D3')
        riqi=ws['B2'].value
        op2_1['G3']=str(riqi)[:11]
        op2_1.merge_cells('G3:H3')
        heji=ws['J2'].value
        op2_1['G19']=heji

        for m in range(1,max_row):
            A=chr(65)
            i='%s%d'%(A,m+4)
            op2_1[i]=m
        B=chr(66)
        C=chr(67)
        D=chr(68)
        E=chr(69)
        F=chr(70)
        for m in range(2,max_row+1):
            mc='%s%d'%(E,m)
            xh='%s%d'%(F,m)
            mc_v=ws[mc].value
            xh_v=ws[xh].value 
            hwmc='%s%d'%(B,m+3)
            hwxh='%s%d'%(C,m+3)
            op2_1[hwmc]=mc_v
            op2_1[hwxh]=xh_v

            dw='%s%d'%(D,m+3)
            if mc_v=='色带' or mc_v=='连接线' :
                op2_1[dw]='条'
            elif mc_v=='标签机' or mc_v=='打印机' or mc_v=='电脑' or mc_v=='高拍仪' :
                op2_1[dw]='台'
            elif mc_v=='墨粉': 
                op2_1[dw]='瓶'
            elif mc_v=='维修费': 
                op2_1[dw]='次'
            else:
                op2_1[dw]='个'
        for m in range(2,max_row+1):
            for n in range(71,74):#chr(97)='a'
                n_1=n-2
                n=chr(n)#ASCII字符
                n_1=chr(n_1)
                i='%s%d'%(n,m)#单元格编号
                i_1='%s%d'%(n_1,m+3)
                cell1=ws[i].value#获取data单元格数据
                op2_1[i_1].value=cell1#赋值到test单元格

    else:
        op2=op.load_workbook('/Users/wangwenyan/Desktop/python/muban_2.xlsx')
        op2_1=op2['Sheet1']

        fphm=ws['C2'].value
        op2_1['C2']=fphm
        op2_1.merge_cells('C2:D2')
        dwmc=ws['D2'].value
        op2_1['C3']=dwmc
        op2_1.merge_cells('C3:D3')
        riqi=ws['B2'].value
        op2_1['G3']=str(riqi)[:11]
        op2_1.merge_cells('G3:H3')
        heji=ws['J2'].value
        op2_1['G19']=heji

        op2_1['C22']=fphm
        op2_1.merge_cells('C22:D22')
        op2_1['C23']=dwmc
        op2_1.merge_cells('C23:D23')
        op2_1['G23']=str(riqi)[:11]
        op2_1.merge_cells('G23:H23')
        op2_1['G39']=heji
        op2_1['A19']='小               计'
        for m in range(1,15):
            A=chr(65)
            i='%s%d'%(A,m+4)
            op2_1[i]=m
        for m in range(15,max_row):
            A=chr(65)
            i='%s%d'%(A,m+10)
            op2_1[i]=m
        B=chr(66)
        C=chr(67)
        D=chr(68)
        E=chr(69)
        F=chr(70)
        for m in range (1,16):
            mc='%s%d'%(E,m)
            xh='%s%d'%(F,m)
            mc_v=ws[mc].value
            xh_v=ws[xh].value 
            hwmc='%s%d'%(B,m+3)
            hwxh='%s%d'%(C,m+3)
            op2_1[hwmc]=mc_v
            op2_1[hwxh]=xh_v
        for m in range (16,max_row+1):
            mc='%s%d'%(E,m)
            xh='%s%d'%(F,m)
            mc_v=ws[mc].value
            xh_v=ws[xh].value 
            hwmc='%s%d'%(B,m+9)
            hwxh='%s%d'%(C,m+9)
            op2_1[hwmc]=mc_v
            op2_1[hwxh]=xh_v
        for m in range (2,max_row+1):
            if m in range(2,16):
                dw_1=m+3
            if m in range(16,max_row+1):
                dw_1=m+9
            dw='%s%d'%(D,dw_1)
            if mc_v=='色带' or mc_v=='连接线':
                op2_1[dw]='个'
            elif mc_v=='标签机' or mc_v=='打印机' or mc_v=='电脑' or mc_v=='高拍仪' :
                op2_1[dw]='台'
            elif mc_v=='墨粉': 
                op2_1[dw]='瓶'
            elif mc_v=='维修费': 
                op2_1[dw]='次'
            else:
                op2_1[dw]='个'
        for m in range (1,16):
            for n in range(71,74):#chr(97)='a'
                n_1=n-2
                n=chr(n)#ASCII字符
                n_1=chr(n_1)
                i='%s%d'%(n,m)#单元格编号
                i_1='%s%d'%(n_1,m+3)
                cell1=ws[i].value#获取data单元格数据
                op2_1[i_1].value=cell1#赋值到test单元格
        for m in range (16,max_row+1):
            for n in range(71,74):#chr(97)='a'
                n_1=n-2
                n=chr(n)#ASCII字符
                n_1=chr(n_1)
                i='%s%d'%(n,m)#单元格编号
                i_1='%s%d'%(n_1,m+9)
                cell1=ws[i].value#获取data单元格数据
                op2_1[i_1].value=cell1#赋值到test单元格
        G=chr(71)
        li1=[]
        li2=[]
        for s in range(5,19):
            je='%s%d'%(G,s)
            li1.append(op2_1[je].value)
        op2_1['G19']=sum(li1)
        for s in range(25,max_row+10):
            je='%s%d'%(G,s)
            li2.append(op2_1[je].value)
            su=sum(li2)
        op2_1['G38']=sum(li2)
        op2_1['A38']='小               计'
        op2_1.merge_cells('A38:F38')
            
    op2.save('/Users/wangwenyan/Desktop/python/输出/输出%s.xlsx'%(fphm))


# 在下方输入需要合并的文件所在文件夹位置
path2='/Users/wangwenyan/Desktop/python/输出'
# 在下方输入合并后Excel的路径和文件名
work=xlsxwriter.Workbook('/Users/wangwenyan/Desktop/python/output.xlsx')


# 新建一个sheet
sheet=work.add_worksheet('combine')

filelist2=os.listdir(path2)
filelist2.sort()
print(len(filelist2))

# Main
file_name='';
x1=1; x2=1;
fileNum = len(filelist2)
print("在该目录下有%d个xlsx文件"%fileNum)

workfomat = work.add_format({
    # 'border':1,
    'font_name': '黑体',  # 字体. 默认值 "Arial"
    'font_size': 14,  # 字号. 默认值 11

})

for file in filelist2:
    if '输出' in file:                       #此处需要修改！
        file_name = os.path.join(path2,file) 
    else:
        continue

    workbook=xlrd.open_workbook(file_name)
    sheet_name=workbook.sheet_names()

    for file_1 in sheet_name:
        table=workbook.sheet_by_name(file_1)
        rows=table.nrows
        clos=table.ncols

        for i in range(rows): 
            sheet.write_row('A'+str(x1),table.row_values(i),workfomat)
            x1+=1

    print('正在合并第%d个文件 '%x2)
    print('已完成 ' + file_name)
    x2 += 1;
A=chr(65)
H=chr(72)

sheet.set_default_row(18)
sheet.set_column(0, 0, 6)
sheet.set_column(1, 1, 13)
sheet.set_column(2, 2, 12.5)
sheet.set_column(3, 3, 7)
sheet.set_column(4, 4, 7)
sheet.set_column(5, 5, 11)
sheet.set_column(6, 6, 12)
sheet.set_column(7, 7, 12)
print("已将%d个文件合并完成"%fileNum)

work.close()


op1=op.load_workbook('/Users/wangwenyan/Desktop/python/output.xlsx')
op1_1=op1['combine']
op1_1.delete_cols(9,12)
A=chr(65)
C=chr(67)
G=chr(71)
max_r=op1_1.max_row


align1 = Alignment(horizontal='center',vertical='center')
align2 = Alignment(horizontal='right',vertical='center')
align3 = Alignment(horizontal='left',vertical='center')
fonthl= Font(name=u'黑体',size=18)
for i in range(1,max_r-18,20):
    op1_1.row_dimensions[i].height = 23
    op1_1.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
    op1_1.merge_cells(start_row=i+1, start_column=1, end_row=i+1, end_column=2)
    op1_1.merge_cells(start_row=i+1, start_column=3, end_row=i+1, end_column=4)
    op1_1.merge_cells(start_row=i+1, start_column=7, end_row=i+1, end_column=8)
    op1_1.merge_cells(start_row=i+2, start_column=1, end_row=i+2, end_column=2)
    op1_1.merge_cells(start_row=i+2, start_column=3, end_row=i+2, end_column=4)
    op1_1.merge_cells(start_row=i+2, start_column=7, end_row=i+2, end_column=8)
    op1_1.merge_cells(start_row=i+18, start_column=1, end_row=i+18, end_column=3)

    hm='%s%d'%(A,i)
    op1_1[hm].font=fonthl
    # op1_1[hm].alignment=align2
    # dwmc_i='%s%d'%(A,i+2)
    # op1_1[dwmc_i].alignment=align2
for m in range(1,max_r+1):
    for n in range(1,9):
        op1_1.cell(m,n).alignment=align1
for i in range(1,max_r-18,20):
    hm='%s%d'%(A,i+1)
    op1_1[hm].alignment=align2
    dwmc_1='%s%d'%(A,i+2)
    op1_1[dwmc_1].alignment=align2
    hm_1='%s%d'%(C,i+1)
    op1_1[hm_1].alignment=align3
    dwmc_2='%s%d'%(C,i+2)
    op1_1[dwmc_2].alignment=align3
    bh='%s%d'%(G,i+1)
    op1_1[bh].alignment=align3
    rq='%s%d'%(G,i+2)
    op1_1[rq].alignment=align3
thin = Side(border_style="thin",color="000000")
border1 = Border(top=thin, left=thin, right=thin, bottom=thin)

for a in range (0,max_r-1,20):
    for b in range (4+a,20+a):
        for c in range(1,9):
            op1_1.cell(b,c).border=border1

print('请输入今天的日期，格式为20210827:')

date_1=input()
li=[]
for g in range(2,max_r-17,20):
    hm_1='%s%d'%(C,g)
    hm_2='%s%d'%(C,g+20)
    dj_1='%s%d'%(G,g)
    hmv=op1_1[hm_1].value
    li.append(hmv)
    li = list(set(li))
    
    for i in li:
        ind = li.index(i)+1
        if ind<10:
            djbh=str(date_1)+'00%s'%(ind)
            op1_1[dj_1]=djbh
        elif 10<=ind<=99:
            djbh=str(date_1)+'0%s'%(ind)
            op1_1[dj_1]=djbh
        else:
            djbh=str(date_1)+'%s'%(ind)
            op1_1[dj_1]=djbh
    
    
        
op1.save('/Users/wangwenyan/Desktop/python/完成.xlsx')




