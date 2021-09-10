# -*- coding = utf-8 -*-
####################README#########################
#2021-07-01
#修复迟到和加班在一起显示为正常，需显示为迟
#修复实际出勤为正常和迟到的汇总（汇总语句添加到最后）
#处理流程说明：
#获取文件的月份信息
#去除不参加考勤人员名单
#首先对几个固定的值进行重新赋值['休息','正常','上班缺卡','下班缺卡','上班迟到','下班早退','旷工']为['0','1','1','1','1','1','0']
#有特殊考勤人员名单额外先进行处理，有休息字段，但是加班打卡的需要计算实际考勤,其他人员休息天计数为0
#针对迟到的人员，进行迟到时间/及频次的统计
#针对假期进行判断，比如调休，发生在上午还是下午，进行不同标记
#合并几个dataframe进行excel输出
#把字符的0和1进行替换，符合要求
###################################################


# 计算时间差
import re
import numpy as np 
import pandas as pd 
#导入openpyxl读取第一个单元格内容(特定单元格内容读取)
from openpyxl import load_workbook 


filename1 = '8月月度汇总.xlsx'
filename2 = '8月打卡时间.xlsx'

#读取第一个单元格内容，确定要处理的月份
wsheet = load_workbook(filename1)['月度汇总']
cell_value = wsheet.cell(1,1).value
#print(cell_value)

#使用正则确定考勤月份
#需定义月度计划考勤天数
def year_month_day(month_str):
    result = re.search(r'(\d+-\d+-\d+).*?(\d+-\d+-\d+).*', month_str)
    return {'year':result.group(1).split('-')[0], 
            'month':result.group(1).split('-')[1],'day':result.group(2)[-2:]}

month_days = year_month_day(cell_value)
# print("%s月份有%s天" %(month_days['month'],month_days['day']))


df = pd.read_excel(filename1,header = 3)
#文件有不规则标题，需要重新命名列名
## 需要对列进行重新赋值命名
'''
['姓名', '考勤组', '部门', '工号', '职位', 'UserId', '出勤天数', '休息天数', '工作时长(分钟)',
       '迟到次数', '迟到时长(分钟)', '严重迟到次数', '严重迟到时长', '旷工迟到次数', '早退次数', '早退时长(分钟)',
       '上班缺卡次数', '下班缺卡次数', '旷工天数', '出差时长', '外出时长', '事假', '调休',
       '病假', '年假', '产假', '陪产假',
       '婚假', '丧假', '哺乳假(小时)', '加班总时长', '工作日(转调休)',
       '休息日(转调休)', '节假日(转调休)', 后面是月份的长度作为列名
'''
#更改列名
col1 = ['Unnamed: 0', 'Unnamed: 1', 'Unnamed: 2', 'Unnamed: 3', 'Unnamed: 4',
   'Unnamed: 5', 'Unnamed: 6', 'Unnamed: 7', 'Unnamed: 8', 'Unnamed: 9',
   'Unnamed: 10', 'Unnamed: 11', 'Unnamed: 12', 'Unnamed: 13',
   'Unnamed: 14', 'Unnamed: 15', 'Unnamed: 16', 'Unnamed: 17',
   'Unnamed: 18', 'Unnamed: 19', 'Unnamed: 20', 'Unnamed: 30']
col2 = ['姓名', '考勤组', '部门', '工号', '职位', 'UserId', '出勤天数', '休息天数', '工作时长(分钟)',
   '迟到次数', '迟到时长(分钟)', '严重迟到次数', '严重迟到时长', '旷工迟到次数', '早退次数', '早退时长(分钟)',
       '上班缺卡次数', '下班缺卡次数', '旷工天数', '出差时长', '外出时长','加班总时长']
rename_dic = {}
for col in range(len(col1)):
	rename_dic[col1[col]]= col2[col]
df = df.rename(columns=rename_dic)

# print(df.columns)
#使用UserId作为新的索引
df = df.set_index('UserId')

###### #去除不考勤人员列表
# names = ['金南贝','汪松玉','黄益峰']
# #根据名称获得UserId，根据UserId来删除行

def get_id(name):
	return df[df['姓名'] == name].index.tolist()[0]
# for name in names:
###### 	df = df.drop(get_id(name))

#处理考勤数据
#针对日期数据进行处理，df取姓名和日期数据部分
df_data = df.iloc[:, 33:(len(df.columns))]
df_data.fillna('0',inplace = True)

#更改列名
col1 = df_data.columns.tolist()
col2 = list(range(1,len(df_data.columns)+1))
rename_dic = {}
for col in range(len(col1)):
	rename_dic[col1[col]] = col2[col]
df_data = df_data.rename(columns = rename_dic)
# print((df_data.columns))

###########################
#准备数据给统计迟到，请假事件等使用 #
###########################
df_cp = df_data.copy()

def df_event_day(event_name):
	df_later = df_cp.copy()
	for col in df_later.columns:
		df_later[col] = df_later[col].str.contains(event_name,na=False)
	return df_later

###########################

#将"休息"天也赋值为0，将其他不同的值也赋值不同的值
# old_value = ['休息','正常','上班缺卡','下班缺卡','上班迟到','下班早退','旷工']
# new_value = ['休','1','1','1','迟','1','0']   #后续实际出勤需要加上迟到的次数
old_value = ['休息','正常','上班缺卡','下班缺卡','上班迟到','旷工','下班早退']
new_value = ['休','1','1','1','迟','旷','退']   #后续实际出勤需要加上迟到的次数
df_data.replace(old_value,new_value,inplace = True)

#定义规则除外人员名单,休息日出勤或者出差计算工作日，赋值1
#因为是使用UseId作为索引，因此需要找出这些人的UserId
'''
胡修波:173308126032368469
包志斌:15965056587103610
蒋亚君:195357363733253004
郭秀玲:1239576338-1101030394
李敏:01270616845761
黄文亮:0437252639884747
'''
# exception_userids = ['173308126032368469','15965056587103610','195357363733253004','1239576338-1101030394','01270616845761','0437252639884747']
exception_names = ['包志斌','蒋亚君','郭秀玲','李敏']

for row in exception_names:
	# index = df_data[get_id(row)]
	# print(index)
	df_data.loc[get_id(row)][df_data.loc[get_id(row)].str.contains('出差|外勤|打卡',na=False)] = '1'


#针对其他员工进行处理的出差外勤，加班等进行处理  休息|年假|调休|病假|事假|婚假|丧假|产假|陪产假|哺乳假

# print(df_data.loc['173308126032368469'])

#所有员工的行.
for each_row in df_data.index.values:
	df_data.loc[each_row][df_data.loc[each_row].str.contains('休息', na=False)] = '0'
	# df_data.loc[each_row][df_data.loc[each_row].str.contains('休息|年假|调休|病假|事假|婚假|丧假|产假|陪产假|哺乳假', na=False)] = '0'
	# NewAdd
	df_data.loc[each_row][df_data.loc[each_row].str.contains('上班迟到', na=False)] = '迟'
	df_data.loc[each_row][df_data.loc[each_row].str.contains('出差|外出|加班|外勤', na=False)] = '1' 



# print(df_data.loc['沈斐'][df_data.loc['沈斐'] == '1'].value_counts())
# print(df_data.loc['0956483708943925'])


#处理迟到时间次数统计问题
#只对标记上班知道的日期进行处理
#读取另外一个excel表格

df2 = pd.read_excel(filename2,header = 2)
df2 = df2.set_index('UserId')
df2 = df2.iloc[:,5:(len(df.columns))]
# print(df2.columns)
df2.columns = df_cp.columns


#迟到的单元格
df2 = df2[df_event_day('上班迟到')]


# 针对多次重复打卡，取第一次和最后一次打卡时间，返回是元组
def parse_time(timestring):
	if isinstance(timestring,str):
		time_lst = timestring.split('\n')
		return time_lst[0].strip(), time_lst[len(time_lst)-1].strip()
df2 = df2.applymap(parse_time)

check_in = "09:00"
def later_minutes_count(cell):
	# check_out = "17:30"
	# 打卡时间表内已经只显示迟到的打卡时间
	if isinstance(cell,tuple):
		dt_delta1 = pd.to_datetime(cell[0]) - pd.to_datetime(check_in)
		# dt_delta2 = pd.to_datetime(check_out) - pd.to_datetime(cell[1])
		if 0 < dt_delta1.total_seconds()/60 < 30:
			return "迟到小于30分钟"
		elif 30 < dt_delta1.total_seconds()/60 < 60:
			return "迟到小于60分钟"
		else:
			return "迟到大于60分钟"
df2 = df2.applymap(later_minutes_count)

# print((df2 == '迟到小于30分钟').sum(axis = 1))
df2['迟到小于30分钟'] = (df2 == '迟到小于30分钟').sum(axis = 1)
df2['迟到小于60分钟'] = (df2 == '迟到小于60分钟').sum(axis = 1)
df2['迟到大于60分钟'] = (df2 == '迟到大于60分钟').sum(axis = 1)



#判断请假是在早上还是下午，使用正则找到调休字段，用于判断调休时间


def time_judgement(cell):
	#newAdd 2021-07-01 处理迟到|事件(调休)
	re_pattern = re.compile(r'([^\x00-\xff]+,)*([^\x00-\xff]+)(\d{2}-\d{2} )(\d{2}:\d{2})[^\x00-\xff](\d{2}-\d{2} )(\d{2}:\d{2})') # 编译
	if isinstance(cell, str):
		#判断str是否含有迟到，如果有迟到
		right_mark = '√'
		if "迟到" in cell:
			right_mark = '迟'

		cell = cell.split(',')
		t = 0
		event_start = event_end =  []
		# print(cell)
		for each_value in cell:
			if re_pattern.match(each_value):
				group_len = len(re_pattern.match(each_value).groups())
				t1 = re_pattern.match(each_value).groups()[group_len - 3]
				event_start.append(t1)
				t2 = re_pattern.match(each_value).groups()[group_len - 1]
				event_end.append(t2)
				event_shortname = re_pattern.match(each_value).groups()[group_len - 5]
				# print(event_shortname)
				t += (pd.to_datetime(t2) - pd.to_datetime(t1)).total_seconds()/3600
		if t >= 8.5:
			return event_shortname[0]
		elif (pd.to_datetime(event_start).min() - pd.to_datetime(check_in)).total_seconds()/3600 < 3:
			return event_shortname[0] + "|" + right_mark
		else:
			return  right_mark + "|" + event_shortname[0]


events_lst = ['调休','年假','病假','事假','婚假','丧假','产假','陪产假','哺乳假']
# df_cp = df_data.copy().drop(columns=['实际出勤'])
for name in events_lst:
	df3 = df_cp[df_event_day(name)]
	df3 = df3.applymap(time_judgement)
	df_data.update(df3)
	df_cp.update(df3)

# print(df_data.loc['0513192160858419'])
# 

#把字符"1"替换为"√",把'0'替换为'休'
df_data.replace(['1','0'],['√','休'], inplace = True)
#增加"实际出勤"列,数值为每行的"1"的个数, 要再加上已经被标记为"迟"的天数
df_data['实际出勤'] = (df_data == '√').sum(axis = 1) + (df_data == '迟').sum(axis = 1)
# print(df_data['实际出勤'])

#拼接新的dataframe
#选择需要的列
'''['姓名', '考勤组', '部门', '工号', '职位', 'UserId', '出勤天数', '休息天数', '工作时长(分钟)',
   '迟到次数', '迟到时长(分钟)', '严重迟到次数', '严重迟到时长', '旷工迟到次数', '早退次数', '早退时长(分钟)',
       '上班缺卡次数', '下班缺卡次数', '旷工天数', '出差时长', '外出时长','加班总时长']
'''
df_p1 = df[['姓名', '部门', '职位', '休息天数', '工作时长(分钟)',
   '迟到次数', '迟到时长(分钟)', '严重迟到次数', '严重迟到时长', '旷工迟到次数', '早退次数', '早退时长(分钟)',
       '上班缺卡次数', '下班缺卡次数', '旷工天数', '出差时长', '外出时长','事假(天)', '调休(天)', '病假(天)',
       '年假(天)', '产假(天)', '陪产假(天)', '婚假(天)', '丧假(天)', '哺乳假(小时)','加班总时长']]
df_p2 = df_data
df_p3 = df2.iloc[:,:]  #测试时合并整个df2,后续只合并后三列统计值df2[:,-3:]
# df_p3 = df2.iloc[:,-3:]
df_final = pd.concat([df_p1,df_p2,df_p3],axis=1)



#保存到excel文件
df_final.to_excel('%s月份考勤处理结果.xlsx' %(month_days['month']), sheet_name='Sheet1', index=False, header=True)
